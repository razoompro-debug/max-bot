from flask import Flask, request, jsonify
import requests
import os
import time

from openpyxl import Workbook, load_workbook
from datetime import datetime

# ==================================================
# НАСТРОЙКИ
# ==================================================

# ТОКЕН MAX BOT
TOKEN = "f9LHodD0cOIloCdM4S4u2QEo0hF1yDOLdVH23RWt5jZwWyIWM82UMhtRYvdd5vPJJ4jYfNEjdPrbnSAV4siW"

# API MAX
API_URL = "https://platform-api.max.ru"

# CHAT ID администратора
ADMIN_CHAT_ID = "290993557"

# Excel файл
EXCEL_FILE = "orders.xlsx"

# Flask
app = Flask(__name__)

# ==================================================
# СОЗДАНИЕ EXCEL
# ==================================================

def create_excel():

    if not os.path.exists(EXCEL_FILE):

        wb = Workbook()
        ws = wb.active

        ws.title = "Orders"

        ws.append([
            "Дата",
            "Chat ID",
            "Имя",
            "Заказ"
        ])

        wb.save(EXCEL_FILE)

        print("EXCEL CREATED")

# ==================================================
# СОХРАНЕНИЕ ЗАКАЗА
# ==================================================

def save_order(chat_id, user_name, order_text):

    try:

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        ws.append([
            datetime.now().strftime("%d.%m.%Y %H:%M"),
            str(chat_id),
            user_name,
            order_text
        ])

        wb.save(EXCEL_FILE)

        print("ORDER SAVED")

    except Exception as e:

        print("SAVE ORDER ERROR:", str(e))

# ==================================================
# ОТПРАВКА СООБЩЕНИЯ
# ==================================================

def send_message(chat_id, text):

    try:

        # антиспам
        time.sleep(2)

        url = f"{API_URL}/messages"

        headers = {
            "Authorization": f"Bearer {TOKEN}",
            "Content-Type": "application/json"
        }

        payload = {
            "recipient": {
                "chat_id": str(chat_id)
            },
            "body": {
                "text": str(text)
            }
        }

        response = requests.post(
            url,
            headers=headers,
            json=payload,
            timeout=15
        )

        print("===================================")
        print("SEND MESSAGE")
        print("CHAT:", chat_id)
        print("STATUS:", response.status_code)
        print("RESPONSE:", response.text)
        print("===================================")

        # если лимит
        if response.status_code == 429:

            print("LIMIT 429 -> WAIT 5 SEC")

            time.sleep(5)

        return response

    except Exception as e:

        print("SEND ERROR:", str(e))

# ==================================================
# ГЛАВНАЯ СТРАНИЦА
# ==================================================

@app.route("/")
def home():

    return "MAX BOT WORKING"

# ==================================================
# WEBHOOK
# ==================================================

@app.route("/webhook", methods=["POST"])
def webhook():

    try:

        data = request.json

        print("WEBHOOK DATA:", data)

        # сообщение
        message = data.get("message", {})

        # текст
        body = message.get("body", {})
        text = body.get("text", "").strip()

        # если текст пустой
        if not text:

            return jsonify({
                "status": "empty_message"
            })

        # пользователь
        sender = message.get("sender", {})
        user_name = sender.get(
            "first_name",
            "Пользователь"
        )

        # чат
        recipient = message.get("recipient", {})
        chat_id = recipient.get("chat_id")

        print("CHAT ID:", chat_id)
        print("TEXT:", text)

        # если нет chat_id
        if not chat_id:

            return jsonify({
                "status": "no_chat_id"
            })

        # ==================================================
        # /start
        # ==================================================

        if text == "/start":

            answer = f"""
👋 Здравствуйте, {user_name}!

Добро пожаловать в бот заказов.

Доступные команды:

/start — запуск
/help — помощь
/order — оформить заказ

Пример заказа:

/order Хочу заказать баннер
"""

            send_message(chat_id, answer)

        # ==================================================
        # /help
        # ==================================================

        elif text == "/help":

            help_text = """
📌 Команды:

/start — запуск бота
/help — помощь
/order — оформить заказ

Пример:

/order Хочу заказать баннер
"""

            send_message(
                chat_id,
                help_text
            )

        # ==================================================
        # /order
        # ==================================================

        elif text.startswith("/order"):

            order_text = text.replace(
                "/order",
                ""
            ).strip()

            # пустой заказ
            if order_text == "":

                send_message(
                    chat_id,
                    "❌ Введите текст заказа.\n\nПример:\n/order Хочу заказать баннер"
                )

            else:

                # сохранить заказ
                save_order(
                    chat_id,
                    user_name,
                    order_text
                )

                # ответ клиенту
                send_message(
                    chat_id,
                    f"""
✅ Заказ принят!

📦 Ваш заказ:

{order_text}
"""
                )

                # сообщение админу
                admin_message = f"""
🆕 Новый заказ

👤 Клиент:
{user_name}

💬 Заказ:
{order_text}

🆔 Chat ID:
{chat_id}
"""

                send_message(
                    ADMIN_CHAT_ID,
                    admin_message
                )

        # ==================================================
        # НЕИЗВЕСТНАЯ КОМАНДА
        # ==================================================

        else:

            send_message(
                chat_id,
                "❌ Неизвестная команда.\nНапишите /help"
            )

        return jsonify({
            "status": "ok"
        })

    except Exception as e:

        print("WEBHOOK ERROR:", str(e))

        return jsonify({
            "status": "error",
            "message": str(e)
        })

# ==================================================
# ЗАПУСК
# ==================================================

create_excel()

if __name__ == "__main__":

    port = int(os.environ.get("PORT", 10000))

    app.run(
        host="0.0.0.0",
        port=port
    )